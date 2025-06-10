"""
    Script to automate the process of getting cameras with state "not responding".

    Author: Maya Aguirre
    Date: 2023-02-02
"""

# Import libraries
import ast
import subprocess
import os
from dotenv import load_dotenv
import FreeSimpleGUI as Sg
import openpyxl
import sys
from datetime import date


def resource_path(relative_path: str):
    """Get the path of the resource.
       PyInstaller creates a temp folder and stores path in _MEIPASS"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def manage_error(error: str):
    """Manage common errors while running the script."""
    if "No se puede enlazar el argumento al parámetro 'Hardware' porque es nulo." in error or "ServerNotFoundMIPException" in error:
        window_alert("El servidor no responde.")
    elif error.find("ejecución de scripts está deshabilitada") > 0 or error.find("running scripts is disabled on this system") > 0:
        command = os.environ.get("SET_POLICY")
        # Set policy to run script
        result_policy = subprocess.run(["powershell", "-Command", command], capture_output=True,
                                       encoding='cp437')
        if result_policy.stderr != '':
            manage_error(result_policy.stderr)
    else:
        window_alert(error)


def install_module():
    """
        Function to install MilestonePSTools PowerShell module if not exists in the device
        Gets the path from the .env, then tries to run the script, if it gets an error, call the function manage_error()
        Once it finished to set the policy, runs again the script to install the module.
    """
    try:
        # Install the module
        ruta_script = resource_path(os.environ.get("SCRIPT_INSTALL_MODULE"))
        result = subprocess.run(["powershell", "-File", ruta_script], capture_output=True, encoding='cp437')
        # If it gets error
        if result.stderr != '':
            manage_error(result.stderr)
        subprocess.run(["powershell", "-File", ruta_script])
        return None
    except Exception as error:
        window_alert(f'An error occurred during installing module: {error}')
        return -1


def validate_module():
    """
        Validates if the module MilestonePSTools is installed in the current machine with a PowerShell command written
        in the .env file.
        If is not installed, call the function install_module()
    """
    try:
        print('Validating module')
        command = os.environ.get("VALIDATE_MODULE")
        result = subprocess.run(["powershell", "-Command", command], capture_output=True)
        if result.stdout == b'':
            install_module()
            return None
        return None
    except Exception as error:
        window_alert(f'An error occurred during validating module: {error}')
        return -1


def window_alert(message: str):
    """Create a window to alert the user from an error or the process is finished."""
    # Error message
    if "error" in message.lower():
        layout = [[Sg.Text(message, size=(50, 15), justification="center")],
                  [Sg.Button("Ok", border_width=3, size=(7, 1))]]
        window = Sg.Window("Alert", layout, element_justification="center")
        while True:
            event, values = window.read()
            if event == "Ok" or event == Sg.WIN_CLOSED:
                break
        return -1
    # Success message
    else:
        layout = [[Sg.Text(message, size=(50, 4), justification="center")]]
        window = Sg.Window("Alert", layout, element_justification="center")
        while True:
            event, values = window.read(timeout=5000)
            if event == Sg.WIN_CLOSED or event == Sg.TIMEOUT_KEY:
                break
    window.close()
    return None


def clear_response(result: str):
    """Clear the response and set the header."""
    if not result:
        return result
    else:
        result = result.split('\n')
        result.pop(0)
        result.pop(1)
        result[0] = 'CAMARAS'
    return result


def response_to_xlsx(response: str, server: str):
    """
        Save the response to a xlsx file.
        First validate if the file exists, if not create it, if it exists, open it and select the sheet.
        Then write the data to the sheet. Finally, save the workbook.
    """
    try:
        print(f'Formatting response.')
        today_date = date.today() # Get date
        file_name = f'cameras_not_responding-{today_date}.xlsx'
        if not os.path.exists(file_name):
            # Create a workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = server
        else:
            # Open the workbook
            workbook = openpyxl.load_workbook(file_name)
            # Select the sheet
            if server in workbook.sheetnames:
                sheet = workbook[server]
            else:
                sheet = workbook.create_sheet(server)
        # Write data to the workbook
        if not response:
            sheet.cell(row=1, column=1).value = "No hay cámaras sin funcionar"
        else:
            for index in range(len(response)):
                sheet.cell(row=index + 1, column=1).value = response[index].rstrip()
        # Save the workbook
        print(f'Saving workbook.')
        workbook.save(file_name)
        workbook.close()
        return None
    except Exception as e:
        window_alert(f'An error occurred saving the response: {e}')
        return -1


def main():
    """Main processes to get the cameras with state: Not Responding"""
    try:
        # Load the environment variables
        load_env()
        # Validate MilestonePSTools exists
        validate_module()
        # Dictionary with the IP of the servers
        servers = dict(ast.literal_eval(os.environ.get("IP_SERVERS")))
        # Run scripts in the servers
        for server in servers:
            print(server)
            # if server == "GDL-CORONA":
            #     command = os.environ.get("COMMAND_CORONA").format(servers[server])
            # else:
            command = os.environ.get("COMMAND").format(servers[server])
            result = subprocess.run(["powershell", "-Command", command], capture_output=True, encoding="cp437")
            # If there is an error
            if result.returncode != 0 or result.stderr != '':
                manage_error(f'Error occurred during getting response: {result.stderr}')
            # If there is not an error
            else:
                response_to_xlsx(clear_response(str(result.stdout)), str(server))
        window_alert("Process finished")
        return None
    except Exception as e:
        window_alert(f'An error occurred: {e}')
        return -1


def load_env():
    """Load the environment variables."""
    try:
        load_dotenv(resource_path(".\\resources\\.env"))
    except Exception as error:
        window_alert(f'An error occurred loading the environment variables: {error}')
        sys.exit()


if __name__ == '__main__':
    main()
