"""
    Script to automate the process of getting cameras which are not responding.

    Author: Maya Aguirre
    Date: 2023-02-02
"""

# Import libraries
import ast
import subprocess
import os
from dotenv import load_dotenv
import PySimpleGUI as Sg
import openpyxl
import sys


def resource_path(relative_path):
    """Get the path of the resource.
       PyInstaller creates a temp folder and stores path in _MEIPASS"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def manage_error(error):
    """Manage the error."""
    if "No se puede enlazar el argumento al parámetro 'Hardware' porque es nulo." in error:
        window_alert("The server is not responding")
    elif "CommandNotFoundException" in error:
        # Install the module
        ruta_script = resource_path(str(os.environ.get("SCRIPT_INSTALL_MODULE")))
        subprocess.run(["powershell", "-File", ruta_script])
    else:
        window_alert(f"An error occurred: {error}")


def window_alert(message):
    """Create a window to alert the user."""
    # Error message
    if "error" in message.lower():
        layout = [[Sg.Text(message, size=(50, 15), justification="center")],
                  [Sg.Button("Ok", border_width=3, size=(7, 1))]]
        window = Sg.Window("Alert", layout, element_justification="center")
        while True:
            event, values = window.read()
            if event == "Ok" or event == Sg.WIN_CLOSED:
                break
    # Success message
    else:
        layout = [[Sg.Text(message, size=(50, 4), justification="center")]]
        window = Sg.Window("Alert", layout, element_justification="center")
        while True:
            event, values = window.read(timeout=5000)
            if event == Sg.WIN_CLOSED or event == Sg.TIMEOUT_KEY:
                break
    window.close()


def clear_response(result):
    """Clear the response."""
    if not result:
        return result
    else:
        result = result.split('\n')
        result.pop(0)
        result.pop(1)
        result[0] = 'CAMARAS'

    return result


def response_to_xlsx(response, server):
    """Save the response to a xlsx file."""
    try:
        if not os.path.exists("cameras_not_responding.xlsx"):
            # Create a workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = server
        else:
            # Open the workbook
            workbook = openpyxl.load_workbook("cameras_not_responding.xlsx")
            # Select the sheet
            if server in workbook.sheetnames:
                sheet = workbook[server]
            else:
                sheet = workbook.create_sheet(server)
        # Write data to the workbook
        if not response:
            sheet.cell(row=1, column=1).value = "No hay cámaras sin funcionar"
        else:
            for index in range(len(response)):
                sheet.cell(row=index + 1, column=1).value = response[index].rstrip()
        # Save the workbook
        workbook.save("cameras_not_responding.xlsx")
        workbook.close()
    except Exception as e:
        window_alert(f"An error occurred saving the response: {e}")


def main():
    """Main function."""
    try:
        # Load the environment variables
        load_env()
        # Dictionary with the IP of the servers
        IP_SERVERS = ast.literal_eval(os.environ.get("IP_SERVERS"))
        # Run scritps in the servers
        for server in IP_SERVERS:
            if server == "GDL-CORONA":
                command = os.environ.get("COMMAND_CORONA").format(IP_SERVERS[server])
            else:
                command = os.environ.get("COMMAND").format(IP_SERVERS[server])
            result = subprocess.run(["powershell", "-Command", command], capture_output=True, encoding="cp437")
            # If there is an error
            if result.returncode != 0:
                manage_error(result.stderr)
            # If there is not an error
            else:
                response_to_xlsx(clear_response(result.stdout), server)
        window_alert("Process finished successfully")
    except Exception as e:
        window_alert(f"An error occurred: {e}")


def load_env():
    """Load the environment variables."""
    try:
        load_dotenv(resource_path("./resources/.env"))
    except Exception as error:
        window_alert(f"An error occurred loading the environment variables: {error}")
        sys.exit()


if __name__ == '__main__':
    main()
